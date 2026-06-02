import argparse
import csv
import json
import re
from pathlib import Path
from typing import Iterable, List, Sequence

from src.schemas import JsonDict, unique_dicts


SIGNAL_CODE_RE = re.compile(r"\bS_[A-Z0-9_]+\b")
FAULT_CODE_RE = re.compile(r"\bDEM_[A-Z0-9_]+\b")
BRACED_ENTITY_RE = re.compile(r"\{([^}]+)\}")
ACRONYM_RE = re.compile(r"\b[A-Z]{2,}(?:_[A-Z0-9]+)*\b")
DOMAIN_WORDS = (
    "vehicle",
    "driver",
    "column",
    "assist",
    "torque",
    "speed",
    "main",
    "secondary",
    "steering",
    "motor",
    "road",
)
DOMAIN_PHRASE_RE = re.compile(
    rf"\b(?:the\s+)?(?:{'|'.join(DOMAIN_WORDS)})(?:\s+(?:{'|'.join(DOMAIN_WORDS)})){{0,3}}\s+"
    r"(?:speed|torque|demand|fault|indicator|signal|state)\b",
    re.IGNORECASE,
)

DEFAULT_SIGNAL_COLUMNS = ["signal", "signal name", "signal_name", "name", "canonical_name"]
DEFAULT_TEXT_COLUMNS = ["requirement", "raw_text", "text", "description", "需求"]
DEFAULT_ID_COLUMNS = ["requirement_id", "req id", "req_id", "id", "需求编号"]


def humanize_signal_name(signal_name: str) -> str:
    """Convert S_FOO_BAR into a natural-language alias such as foo bar."""

    name = signal_name.strip()
    if name.upper().startswith("S_"):
        name = name[2:]
    return " ".join(part.lower() for part in name.split("_") if part)


def build_signal_dictionary(rows: Sequence[JsonDict], signal_column: str | None = None) -> List[JsonDict]:
    """Build initial signal entities from Excel/CSV rows containing S_* names."""

    entities: List[JsonDict] = []
    for row in rows:
        signal_names = _extract_signal_names(row, signal_column)
        for signal_name in signal_names:
            humanized = humanize_signal_name(signal_name)
            aliases = [signal_name, humanized]
            if not humanized.endswith(" signal"):
                aliases.append(f"{humanized} signal")
            entities.append(
                {
                    "canonical_name": signal_name,
                    "type": "signal",
                    "aliases": _unique_strings(aliases),
                    "members": [],
                    "unit": _get_optional_column(row, ["unit", "units"]),
                    "component": _get_optional_column(row, ["component", "module", "subsystem"]),
                    "description": _get_optional_column(row, ["description", "comment", "备注"]),
                    "source": "signal_excel",
                }
            )
    return unique_dicts(entities, ["canonical_name"])


def extract_alias_candidates(
    requirement_rows: Sequence[JsonDict],
    dictionary: Sequence[JsonDict],
    text_column: str | None = None,
    id_column: str | None = None,
) -> List[JsonDict]:
    """Extract alias candidates from requirement rows and attach canonical suggestions."""

    grouped: dict[str, JsonDict] = {}
    for row_index, row in enumerate(requirement_rows, start=1):
        text = _get_required_text(row, text_column, DEFAULT_TEXT_COLUMNS)
        requirement_id = _get_requirement_id(row, id_column, row_index)
        for mention in _extract_mentions(text, dictionary):
            cleaned = clean_mention(mention)
            if not cleaned:
                continue
            suggestion = suggest_canonical(cleaned, dictionary)
            key = cleaned.lower()
            if key not in grouped:
                grouped[key] = {
                    "mention": cleaned,
                    "suggested_canonical": suggestion["canonical_name"],
                    "type": suggestion["type"],
                    "confidence": suggestion["confidence"],
                    "evidence": [],
                    "status": "pending",
                }
            if requirement_id not in grouped[key]["evidence"]:
                grouped[key]["evidence"].append(requirement_id)
            if suggestion["confidence"] > grouped[key]["confidence"]:
                grouped[key]["suggested_canonical"] = suggestion["canonical_name"]
                grouped[key]["type"] = suggestion["type"]
                grouped[key]["confidence"] = suggestion["confidence"]
    return sorted(grouped.values(), key=lambda item: (item["suggested_canonical"] or "", item["mention"].lower()))


def suggest_canonical(mention: str, dictionary: Sequence[JsonDict]) -> JsonDict:
    """Suggest a canonical entity for a mention using exact aliases and token overlap."""

    mention_norm = _normalize_for_match(mention)
    best = {"canonical_name": None, "type": "unknown", "confidence": 0.0}
    for entity in dictionary:
        canonical = str(entity.get("canonical_name", ""))
        aliases = [canonical] + list(entity.get("aliases", []))
        for alias in aliases:
            alias_norm = _normalize_for_match(str(alias))
            if not alias_norm:
                continue
            if mention_norm == _normalize_for_match(canonical):
                return {"canonical_name": canonical, "type": entity.get("type", "signal"), "confidence": 1.0}
            if mention_norm == alias_norm:
                return {"canonical_name": canonical, "type": entity.get("type", "signal"), "confidence": 0.92}
            score = _token_overlap_score(mention_norm, alias_norm)
            if score > best["confidence"]:
                best = {"canonical_name": canonical, "type": entity.get("type", "signal"), "confidence": round(score, 2)}

    if mention.upper().startswith("DEM_"):
        return {"canonical_name": mention.upper(), "type": "fault", "confidence": 1.0}
    if mention.upper() == mention and len(mention) >= 2:
        return {"canonical_name": mention.upper(), "type": "indicator_or_component", "confidence": 0.5}
    return best


def load_tabular_rows(path: Path, sheet_name: str | None = None) -> List[JsonDict]:
    """Load rows from CSV or XLSX. XLSX requires openpyxl."""

    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            return [dict(row) for row in csv.DictReader(file)]
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_rows(path, sheet_name)
    raise ValueError(f"Unsupported file type: {path.suffix}. Use .xlsx, .xlsm, or .csv.")


def write_dictionary(path: Path, entities: Sequence[JsonDict]) -> None:
    """Write the initial dictionary JSON file."""

    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"version": "initial", "entities": list(entities)}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[JsonDict]) -> None:
    """Write JSONL with non-ASCII characters preserved."""

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


def load_dictionary(path: Path) -> List[JsonDict]:
    """Load a dictionary JSON file that is either a list or {'entities': [...]}."""

    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict) and isinstance(payload.get("entities"), list):
        return payload["entities"]
    raise ValueError(f"Dictionary file must be a list or contain an 'entities' list: {path}")


def load_jsonl(path: Path) -> List[JsonDict]:
    """Load JSONL records, skipping blank lines."""

    rows: List[JsonDict] = []
    with path.open("r", encoding="utf-8-sig") as file:
        for line_number, line in enumerate(file, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSONL at {path}:{line_number}: {exc}") from exc
    return rows


def merge_approved_aliases(
    dictionary: Sequence[JsonDict],
    candidates: Sequence[JsonDict],
    create_missing: bool = False,
) -> tuple[List[JsonDict], JsonDict]:
    """Merge approved alias candidates into dictionary entities.

    Only candidates with status='approved' are merged. Users may correct the
    target with canonical_name; otherwise suggested_canonical is used.
    """

    merged = [dict(entity) for entity in dictionary]
    by_canonical = {str(entity.get("canonical_name")): entity for entity in merged}
    report: JsonDict = {
        "approved_candidates": 0,
        "merged_aliases": 0,
        "created_entities": 0,
        "skipped_candidates": 0,
        "unmatched_candidates": [],
    }

    for candidate in candidates:
        if str(candidate.get("status", "")).lower() != "approved":
            report["skipped_candidates"] += 1
            continue

        report["approved_candidates"] += 1
        mention = clean_mention(str(candidate.get("mention", "")))
        canonical = str(candidate.get("canonical_name") or candidate.get("suggested_canonical") or "").strip()
        if not mention or not canonical:
            report["skipped_candidates"] += 1
            report["unmatched_candidates"].append(candidate)
            continue

        entity = by_canonical.get(canonical)
        if entity is None and create_missing:
            entity = {
                "canonical_name": canonical,
                "type": candidate.get("type", "unknown"),
                "aliases": [canonical],
                "members": [],
                "source": "approved_alias_candidate",
            }
            merged.append(entity)
            by_canonical[canonical] = entity
            report["created_entities"] += 1

        if entity is None:
            report["skipped_candidates"] += 1
            report["unmatched_candidates"].append(candidate)
            continue

        aliases = list(entity.get("aliases", []))
        if not _contains_case_insensitive(aliases, mention):
            aliases.append(mention)
            report["merged_aliases"] += 1
        entity["aliases"] = _unique_strings(aliases)

    return merged, report


def merge_from_files(args: argparse.Namespace) -> JsonDict:
    """Merge approved JSONL candidates into a dictionary JSON file."""

    dictionary = load_dictionary(Path(args.dictionary))
    candidates = load_jsonl(Path(args.candidates))
    merged, report = merge_approved_aliases(dictionary, candidates, create_missing=args.create_missing)
    write_dictionary(Path(args.output), merged)
    if args.report_output:
        Path(args.report_output).parent.mkdir(parents=True, exist_ok=True)
        Path(args.report_output).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def build_from_files(args: argparse.Namespace) -> tuple[List[JsonDict], List[JsonDict]]:
    """Build dictionary and candidates from command-line inputs."""

    signal_rows = load_tabular_rows(Path(args.signal_excel), args.signal_sheet)
    dictionary = build_signal_dictionary(signal_rows, args.signal_column)
    write_dictionary(Path(args.output), dictionary)

    candidates: List[JsonDict] = []
    if args.requirements_excel:
        requirement_rows = load_tabular_rows(Path(args.requirements_excel), args.requirements_sheet)
        candidates = extract_alias_candidates(
            requirement_rows,
            dictionary,
            text_column=args.requirement_text_column,
            id_column=args.requirement_id_column,
        )
        write_jsonl(Path(args.candidates_output), candidates)
    return dictionary, candidates


def build_arg_parser() -> argparse.ArgumentParser:
    """Create CLI parser for building entity dictionaries from Excel files."""

    parser = argparse.ArgumentParser(description="Build an initial entity dictionary from signal and requirement Excel files.")
    parser.add_argument("--signal-excel", required=True, help="Path to the signal list .xlsx/.xlsm/.csv file.")
    parser.add_argument("--signal-sheet", default=None, help="Optional signal workbook sheet name.")
    parser.add_argument("--signal-column", default=None, help="Optional column containing S_* signal names.")
    parser.add_argument("--requirements-excel", default=None, help="Optional requirement .xlsx/.xlsm/.csv file.")
    parser.add_argument("--requirements-sheet", default=None, help="Optional requirements workbook sheet name.")
    parser.add_argument("--requirement-text-column", default=None, help="Optional requirement text column name.")
    parser.add_argument("--requirement-id-column", default=None, help="Optional requirement ID column name.")
    parser.add_argument("--output", default="config/entities/signals.json", help="Output dictionary JSON path.")
    parser.add_argument(
        "--candidates-output",
        default="data/entity_alias_candidates.jsonl",
        help="Output alias candidate JSONL path.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    """CLI entry point."""

    parser = build_arg_parser()
    args = parser.parse_args(argv)
    dictionary, candidates = build_from_files(args)
    print(f"Wrote {len(dictionary)} dictionary entities to {args.output}")
    if args.requirements_excel:
        print(f"Wrote {len(candidates)} alias candidates to {args.candidates_output}")
    return 0


def _load_xlsx_rows(path: Path, sheet_name: str | None) -> List[JsonDict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Reading .xlsx files requires openpyxl. Install it or use a .csv input file.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: List[JsonDict] = []
    for values in rows[1:]:
        if not any(value is not None and str(value).strip() for value in values):
            continue
        result.append({headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]})
    return result


def _extract_signal_names(row: JsonDict, signal_column: str | None) -> List[str]:
    if signal_column:
        return _find_signal_codes(str(row.get(signal_column, "")))
    inferred_column = _find_column(row, DEFAULT_SIGNAL_COLUMNS)
    if inferred_column:
        return _find_signal_codes(str(row.get(inferred_column, "")))
    names: List[str] = []
    for value in row.values():
        names.extend(_find_signal_codes(str(value)))
    return _unique_strings(names)


def _extract_mentions(text: str, dictionary: Sequence[JsonDict]) -> List[str]:
    mentions: List[str] = []
    mentions.extend(match.group(0) for match in SIGNAL_CODE_RE.finditer(text))
    mentions.extend(match.group(0) for match in FAULT_CODE_RE.finditer(text))
    mentions.extend(match.group(1) for match in BRACED_ENTITY_RE.finditer(text))
    mentions.extend(match.group(0) for match in ACRONYM_RE.finditer(text) if not match.group(0).startswith(("S_", "DEM_")))
    mentions.extend(match.group(0) for match in DOMAIN_PHRASE_RE.finditer(text))

    for entity in dictionary:
        for alias in entity.get("aliases", []):
            alias_text = str(alias)
            if len(alias_text) < 3:
                continue
            pattern = re.compile(rf"\b{re.escape(alias_text)}\b", re.IGNORECASE)
            mentions.extend(match.group(0) for match in pattern.finditer(text))
    return _unique_strings(mentions)


def clean_mention(mention: str) -> str:
    """Remove leading fillers and normalize whitespace while preserving display case."""

    cleaned = re.sub(r"\s+", " ", mention.strip(" .,;:()[]{}")).strip()
    cleaned = re.sub(r"^(?:the|a|an|with increasing|with decreasing)\s+", "", cleaned, flags=re.IGNORECASE)
    return cleaned.strip()


def _get_required_text(row: JsonDict, column: str | None, defaults: Sequence[str]) -> str:
    selected = column or _find_column(row, defaults)
    if selected:
        return str(row.get(selected, "") or "")
    return " ".join(str(value) for value in row.values() if value is not None)


def _get_requirement_id(row: JsonDict, column: str | None, row_index: int) -> str:
    selected = column or _find_column(row, DEFAULT_ID_COLUMNS)
    if selected and row.get(selected):
        return str(row[selected])
    return f"ROW_{row_index}"


def _get_optional_column(row: JsonDict, candidates: Sequence[str]) -> str:
    column = _find_column(row, candidates)
    value = row.get(column, "") if column else ""
    return "" if value is None else str(value)


def _find_column(row: JsonDict, candidates: Sequence[str]) -> str | None:
    lowered = {str(key).strip().lower(): key for key in row.keys()}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    for key in row.keys():
        key_norm = str(key).strip().lower()
        if any(candidate.lower() in key_norm for candidate in candidates):
            return str(key)
    return None


def _find_signal_codes(value: str) -> List[str]:
    return _unique_strings(match.group(0) for match in SIGNAL_CODE_RE.finditer(value))


def _unique_strings(values: Iterable[str]) -> List[str]:
    seen = set()
    result = []
    for value in values:
        if not value:
            continue
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _contains_case_insensitive(values: Iterable[str], target: str) -> bool:
    target_key = target.lower()
    return any(str(value).lower() == target_key for value in values)


def _normalize_for_match(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("_", " ").strip().lower())


def _token_overlap_score(left: str, right: str) -> float:
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    if not left_tokens or not right_tokens:
        return 0.0
    overlap = len(left_tokens & right_tokens)
    return overlap / max(len(left_tokens), len(right_tokens))
